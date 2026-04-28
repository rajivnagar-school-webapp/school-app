
import base64
import io
import os
from datetime import date
from excel_generators import generate_gun_slip, generate_parinam
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from supabase import create_client, Client
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'shala-secret-2025')

SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

_FONT_DIR       = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'fonts')
_FONT_REG_PATH  = os.path.join(_FONT_DIR, 'NotoSansGujarati-Regular.ttf')
_FONT_BOLD_PATH = os.path.join(_FONT_DIR, 'NotoSansGujarati-Bold.ttf')

if not os.path.exists(_FONT_REG_PATH):
    _FONT_REG_PATH  = '/usr/share/fonts/truetype/freefont/FreeSerif.ttf'
    _FONT_BOLD_PATH = '/usr/share/fonts/truetype/freefont/FreeSerifBold.ttf'


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def logged_in():    return 'user_id' in session
def is_teacher():   return session.get('role') == 'teacher'
def is_principal(): return session.get('role') == 'principal'

def get_grade(total, out_of=60):
    if out_of == 0: return '-'
    p = (total / out_of) * 100
    if p >= 90:   return 'A+'
    elif p >= 75: return 'A'
    elif p >= 60: return 'B'
    elif p >= 45: return 'C'
    else:         return 'D'

def upload_doc_base64(student_id, doc_name, file_obj):
    if not file_obj or not file_obj.filename: return
    file_bytes = file_obj.read()
    if not file_bytes: return
    ext      = file_obj.filename.rsplit('.', 1)[-1].lower() if '.' in file_obj.filename else 'bin'
    b64_data = base64.b64encode(file_bytes).decode('utf-8')
    doc_url  = f"data:{file_obj.content_type};base64,{b64_data}"
    supabase.table('student_documents').insert({
        'student_id': student_id,
        'doc_name':   doc_name,
        'doc_url':    doc_url,
        'file_ext':   ext
    }).execute()

def get_students_with_att(class_name=None):
    if class_name:
        res = supabase.table('students').select('*').ilike('class', class_name).order('roll_no').execute()
    else:
        res = supabase.table('students').select('*').order('class').execute()
    students = res.data or []
    all_att  = (supabase.table('attendance').select('*').execute().data) or []
    for s in students:
        s_att = [a for a in all_att if a['student_id'] == s['id']]
        s['att_pct'] = round(len([a for a in s_att if a['status'] == 'P']) / len(s_att) * 100) if s_att else 100
    return students

def get_student_full(student_id):
    res = supabase.table('students').select('*').eq('id', student_id).execute()
    if not res.data: return None
    s    = res.data[0]
    docs = supabase.table('student_documents').select('*').eq('student_id', student_id).execute()
    s['documents'] = docs.data or []
    return s

def collect_student_form():
    return {
        'name':                   request.form.get('name', '').strip(),
        'surname':                request.form.get('surname', '').strip(),
        'father_name':            request.form.get('father_name', '').strip(),
        'mother_name':            request.form.get('mother_name', '').strip(),
        'roll_no':                request.form.get('roll_no', '').strip(),
        'gr_number':              request.form.get('gr_number', '').strip(),
        'dob':                    request.form.get('dob', ''),
        'gender':                 request.form.get('gender', ''),
        'caste':                  request.form.get('caste', ''),
        'address':                request.form.get('address', '').strip(),
        'parent_contact':         request.form.get('parent_contact', '').strip(),
        'aadhaar_number':         request.form.get('aadhaar_number', '').strip(),
        'bank_account':           request.form.get('bank_account', '').strip(),
        'bank_name':              request.form.get('bank_name', '').strip(),
        'ifsc_code':              request.form.get('ifsc_code', '').strip(),
        'section':                request.form.get('section', '').strip(),
        'attendance_register_no': request.form.get('attendance_register_no', '').strip(),
    }


# ─────────────────────────────────────────────
# GENERAL ROUTES
# ─────────────────────────────────────────────

@app.route('/')
def index():
    if not logged_in(): return redirect(url_for('login'))
    return redirect(url_for('principal_dashboard') if is_principal() else url_for('teacher_dashboard'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username or not password:
            error = 'Username aur password dono bharo.'
            return render_template('login.html', error=error)
        res = supabase.table('users').select('*').eq('username', username).execute()
        if res.data:
            user = res.data[0]
            if check_password_hash(user['password_hash'], password):
                session.update({'user_id': user['id'], 'name': user['name'],
                                'username': user['username'], 'role': user['role'],
                                'class': user.get('class_assigned') or ''})
                return redirect(url_for('principal_dashboard') if user['role'] == 'principal' else url_for('teacher_dashboard'))
            error = 'Password galat hai.'
        else:
            error = 'Username nahi mila.'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['POST'])
def change_password():
    if not logged_in(): return jsonify({'success': False, 'msg': 'Login required.'})
    old = request.form.get('old_password', '').strip()
    new = request.form.get('new_password', '').strip()
    if not old or not new: return jsonify({'success': False, 'msg': 'Dono fields bharo.'})
    if len(new) < 4: return jsonify({'success': False, 'msg': 'Password 4+ characters ka ho.'})
    user = supabase.table('users').select('*').eq('id', session['user_id']).execute()
    if not user.data: return jsonify({'success': False, 'msg': 'User nahi mila.'})
    if check_password_hash(user.data[0]['password_hash'], old):
        supabase.table('users').update({'password_hash': generate_password_hash(new)}).eq('id', session['user_id']).execute()
        return jsonify({'success': True, 'msg': 'Password change ho gaya!'})
    return jsonify({'success': False, 'msg': 'Purana password galat hai.'})


# ─────────────────────────────────────────────
# SHARED LOGIC
# ─────────────────────────────────────────────

def _add_student_logic(class_name, redirect_to):
    data = collect_student_form()
    data['class'] = class_name
    res = supabase.table('students').insert(data).execute()
    if res.data:
        sid = res.data[0]['id']
        for field, doc_name in [('aadhaar_doc', 'Aadhaar Card'), ('ration_doc', 'Ration Card')]:
            if field in request.files:
                try: upload_doc_base64(sid, doc_name, request.files[field])
                except Exception as e: print(f"[DOC ERROR] {field}:", e)
        for i, f in enumerate(request.files.getlist('custom_doc_file')):
            if f and f.filename:
                names = request.form.getlist('custom_doc_name')
                name  = names[i].strip() if i < len(names) and names[i].strip() else f'Document {i+1}'
                try: upload_doc_base64(sid, name, f)
                except Exception as e: print(f"[CUSTOM DOC ERROR]:", e)
    return redirect(url_for(redirect_to))


def _edit_student_logic(redirect_to):
    sid  = request.form.get('student_id')
    data = collect_student_form()
    data.pop('class', None)
    supabase.table('students').update(data).eq('id', sid).execute()
    doc_name = request.form.get('new_doc_name', '').strip()
    f = request.files.get('new_doc_file')
    if f and f.filename and doc_name:
        try: upload_doc_base64(sid, doc_name, f)
        except Exception as e: print("[EDIT DOC ERROR]:", e)
    return redirect(url_for(redirect_to))


def _save_attendance_data(data):
    today = str(date.today())
    for student_id, status in data.get('attendance', {}).items():
        if status not in ('P', 'A'): continue
        existing = supabase.table('attendance').select('id').eq('student_id', student_id).eq('date', today).execute()
        if existing.data:
            supabase.table('attendance').update({'status': status}).eq('student_id', student_id).eq('date', today).execute()
        else:
            supabase.table('attendance').insert({'student_id': student_id, 'date': today,
                                                 'status': status, 'marked_by': session['user_id']}).execute()


def _save_marks_data(data):
    subject_id = data.get('subject_id')
    semester   = data.get('semester')
    for student_id, vals in data.get('marks', {}).items():
        record = {'student_id': student_id, 'subject_id': subject_id,
                  'semester': semester, 'written_marks': int(vals.get('written', 0)),
                  'participation_marks': int(vals.get('participation', 0)),
                  'academic_year': '2025-26', 'entered_by': session['user_id']}
        existing = supabase.table('marks').select('id').eq('student_id', student_id)\
            .eq('subject_id', subject_id).eq('semester', semester).execute()
        if existing.data:
            supabase.table('marks').update(record).eq('id', existing.data[0]['id']).execute()
        else:
            supabase.table('marks').insert(record).execute()


# ─────────────────────────────────────────────
# PDF GENERATION — WeasyPrint (proper Gujarati)
# ─────────────────────────────────────────────

def _generate_student_pdf(student_id):
    s_res = supabase.table('students').select('*').eq('id', student_id).execute()
    if not s_res.data: return 'Student not found', 404
    s = s_res.data[0]

    att_res   = (supabase.table('attendance').select('*').eq('student_id', student_id).execute().data) or []
    total_d   = len(att_res)
    pres_d    = len([a for a in att_res if a['status'] == 'P'])
    att_pct   = round(pres_d / total_d * 100) if total_d > 0 else 0
    docs_res  = (supabase.table('student_documents').select('*').eq('student_id', student_id).execute().data) or []
    marks_res = (supabase.table('marks')
                 .select('*, subjects(subject_name, max_written, max_participation)')
                 .eq('student_id', student_id).execute().data) or []

    cls_num = s.get('class', '').replace('Class ', '')
    dob_fmt = s.get('dob') or '-'
    if dob_fmt and '-' in str(dob_fmt):
        parts = str(dob_fmt).split('-')
        if len(parts) == 3: dob_fmt = f"{parts[2]}/{parts[1]}/{parts[0]}"

    # ── Build marks rows ──
    marks_rows_p1 = ''
    marks_rows_p3 = ''
    grand_total   = 0
    grand_max     = 0

    for i, m in enumerate(marks_res):
        subj_data  = m.get('subjects') or {}
        subj_name  = subj_data.get('subject_name', '-')
        max_w      = subj_data.get('max_written', 40) or 40
        max_p      = subj_data.get('max_participation', 20) or 20
        max_total  = max_w + max_p
        pass_marks = int(max_total * 0.8)
        tot        = m['written_marks'] + m['participation_marks']
        grade      = get_grade(tot, max_total)
        grand_total += tot
        grand_max   += max_total

        marks_rows_p1 += f"""
        <tr>
          <td class="tl">{subj_name}</td>
          <td>Sem {m['semester']}</td>
          <td>{m['written_marks']}</td>
          <td>{m['participation_marks']}</td>
          <td>{tot}</td>
          <td>{grade}</td>
        </tr>"""

        marks_rows_p3 += f"""
        <tr>
          <td>{i+1}</td>
          <td class="tl">{subj_name}</td>
          <td>{max_total}</td>
          <td>{tot}</td>
          <td>{grade}</td>
          <td></td>
        </tr>"""

    overall_grade    = get_grade(grand_total, grand_max) if grand_max > 0 else '-'
    overall_pct      = round(grand_total / grand_max * 100, 1) if grand_max > 0 else 0
    grand_pass_marks = int(grand_max * 0.8)
    row_count        = len(marks_res)

    # ── Documents — sirf naam ──
    docs_section = ''
    if docs_res:
        doc_items = ''.join(f'<li>• {d["doc_name"]}</li>' for d in docs_res)
        docs_section = f'<div class="sec-title" style="margin-top:12px;">Documents Uploaded</div><ul class="doc-ul">{doc_items}</ul>'

    font_reg  = _FONT_REG_PATH.replace('\\', '/')
    font_bold = _FONT_BOLD_PATH.replace('\\', '/')

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  @font-face {{
    font-family:'GujFont';
    src:url('file://{font_reg}');
    font-weight:normal;
  }}
  @font-face {{
    font-family:'GujFont';
    src:url('file://{font_bold}');
    font-weight:bold;
  }}

  @page {{
    size: A4;
    margin: 0;
  }}

  * {{
    font-family:'GujFont', sans-serif;
    margin:0; padding:0;
    box-sizing:border-box;
  }}

  /* ── Each page is exactly A4 ── */
  .page {{
    width: 210mm;
    height: 297mm;
    overflow: hidden;
    position: relative;
    page-break-after: always;
  }}
  .page:last-child {{
    page-break-after: avoid;
  }}

  /* ══════════════════ PAGE 1 ══════════════════ */
  .p1 {{
    padding: 0;
    display: flex;
    flex-direction: column;
    height: 297mm;
  }}
  .p1-header {{
    background: #1a4f8a;
    color: white;
    padding: 18px 28px;
    flex-shrink: 0;
  }}
  .p1-header h1 {{ font-size:20px; font-weight:bold; }}
  .p1-header p  {{ font-size:12px; color:#cbd5e1; margin-top:4px; }}
  .p1-body {{
    padding: 18px 28px 16px;
    flex: 1;
    display: flex;
    flex-direction: column;
  }}
  .p1-name {{ font-size:17px; font-weight:bold; margin-bottom:3px; color:#1e293b; }}
  .p1-sub  {{ font-size:11px; color:#64748b; margin-bottom:12px; }}
  .divider {{ border:none; border-top:1px solid #e2e8f0; margin:11px 0; }}
  .sec-title {{ font-size:13px; font-weight:bold; margin-bottom:7px; color:#1e293b; }}
  .det-tbl {{ width:100%; border-collapse:collapse; margin-bottom:10px; font-size:11px; }}
  .det-tbl td {{ padding:3.5px 6px; color:#1e293b; }}
  .det-tbl td:first-child {{ color:#64748b; width:145px; }}
  .att-row {{ display:flex; gap:36px; font-size:11px; color:#64748b; margin-bottom:10px; }}
  .att-row b {{ color:#1e293b; }}
  .mk-tbl {{ width:100%; border-collapse:collapse; font-size:11px; }}
  .mk-tbl th {{ background:#1a4f8a; color:white; padding:6px 8px; text-align:center; }}
  .mk-tbl th.tl {{ text-align:left; padding-left:10px; }}
  .mk-tbl td {{ border:1px solid #e2e8f0; padding:5px 8px; text-align:center; background:#f8fafc; color:#1e293b; }}
  .mk-tbl td.tl {{ text-align:left; padding-left:10px; }}
  .doc-ul {{ list-style:none; font-size:11px; color:#1a4f8a; line-height:1.9; margin-top:4px; }}
  .p1-footer {{
    font-size:9px; color:#94a3b8;
    display:flex; justify-content:space-between;
    padding: 10px 28px 14px;
    border-top: 1px solid #e2e8f0;
    flex-shrink: 0;
  }}

  /* ══════════════════ PAGE 2 ══════════════════ */
  .p2 {{
    width: 210mm;
    height: 297mm;
    padding: 10mm;
    display: flex;
    flex-direction: column;
  }}
  .p2-inner {{
    border: 1.5px solid black;
    flex: 1;
    padding: 16px 28px 20px;
    display: flex;
    flex-direction: column;
  }}
  .p2-title {{ text-align:center; font-weight:bold; font-size:16px; margin-bottom:3px; }}
  .p2-sub   {{ text-align:center; font-size:12px; margin-bottom:2px; }}
  .p2-dotline {{
    border:none; border-bottom:2px dotted #666;
    width:200px; margin:4px auto 6px; display:block;
  }}
  .p2-main-title {{
    text-align:center; font-weight:bold; font-size:20px;
    text-decoration:underline; margin:10px 0 4px;
  }}
  .p2-f   {{ text-align:center; font-weight:bold; font-size:13px; margin-top:4px; }}
  .p2-progress {{ text-align:center; font-size:12px; margin-bottom:14px; }}
  .p2-row {{
    display:flex; align-items:baseline;
    margin:8px 0; font-size:12px;
  }}
  .p2-lbl {{ min-width:145px; color:#222; }}
  .p2-val {{
    font-weight:bold;
    border-bottom:1px dotted #666;
    flex:1; padding:0 4px 1px;
    min-height:17px;
  }}
  .p2-name-row {{ display:flex; gap:10px; margin:10px 0 4px; }}
  .p2-name-col {{ flex:1; }}
  .p2-nval {{
    font-weight:bold;
    border-bottom:1px dotted #666;
    font-size:12px; padding-bottom:2px;
    min-height:17px; display:block;
  }}
  .p2-nlbl {{ font-size:9px; text-align:center; color:#555; margin-top:3px; }}
  .p2-divider {{ border:none; border-top:1px solid #444; margin:16px 0 10px; }}
  .p2-bank {{ font-size:12px; line-height:2.2; }}
  .p2-spacer {{ flex:1; }}
  .p2-sigs {{
    display:flex; justify-content:space-between;
    font-size:15px; padding-top:16px;
  }}
  .p2-sig {{ text-align:center; }}
  .p2-sigline {{ border-top:1px solid #333; width:160px; margin:0 auto 5px; }}

  /* ══════════════════ PAGE 3 ══════════════════ */
  .p3 {{
    width: 210mm;
    height: 297mm;
    padding: 10mm;
    display: flex;
    flex-direction: column;
  }}
  .p3-inner {{
    border: 1.5px solid black;
    flex: 1;
    padding: 16px 28px 20px;
    display: flex;
    flex-direction: column;
  }}
  .p3-title {{ text-align:center; font-weight:bold; font-size:16px; margin-bottom:18px; }}
  .p3-info  {{ font-size:11px; margin-bottom:14px; line-height:1.9; }}
  .p3-info-val {{
    font-weight:bold;
    border-bottom:1px solid #333;
    padding:0 4px 1px;
    display:inline-block;
  }}
  .guj-tbl {{ width:100%; border-collapse:collapse; font-size:11px; }}
  .guj-tbl th {{
    background:#ebebeb; border:1px solid #333;
    padding:7px 8px; text-align:center; font-weight:bold;
  }}
  .guj-tbl th.tl {{ text-align:left; }}
  .guj-tbl td {{ border:1px solid #333; padding:6px 8px; text-align:center; }}
  .guj-tbl td.tl {{ text-align:left; }}
  .guj-tbl tr.tot-row td {{ background:#f0f0f0; font-weight:bold; }}
  .p3-spacer {{ flex:1; }}
  .p3-bottom {{
    font-size:15px; display:flex;
    gap:30px; margin-top:16px; margin-bottom:16px;
  }}
  .p3-bottom-val {{
    border-bottom:1px solid #333;
    min-width:110px; display:inline-block;
  }}
  .p3-sigs {{
    display:flex; justify-content:space-between;
    font-size:15px; margin-top:30px;
  }}
  .p3-sig {{ text-align:center; }}
  .p3-sigline {{ border-top:1px solid #333; width:170px; margin:0 auto 5px; }}
</style>
</head>
<body>

<!-- ══════════════ PAGE 1: English Profile ══════════════ -->
<div class="page p1">
  <div class="p1-header">
    <h1>Rajivnagar Primary School</h1>
    <p>Student Profile Report</p>
  </div>
  <div class="p1-body">
    <div class="p1-name">{s.get('name','')} {s.get('surname','')}</div>
    <div class="p1-sub">
      Class: {s.get('class','-')} &nbsp;|&nbsp;
      Roll No: {s.get('roll_no','-')} &nbsp;|&nbsp;
      GR: {s.get('gr_number','-')}
    </div>
    <hr class="divider">

    <div class="sec-title">Student Details</div>
    <table class="det-tbl">
      <tr><td>Date of Birth</td><td>{s.get('dob','-')}</td></tr>
      <tr><td>Gender</td><td>{s.get('gender','-')}</td></tr>
      <tr><td>Caste</td><td>{s.get('caste','-')}</td></tr>
      <tr><td>Father Name</td><td>{s.get('father_name','-')}</td></tr>
      <tr><td>Mother Name</td><td>{s.get('mother_name','-')}</td></tr>
      <tr><td>Address</td><td>{s.get('address','-')}</td></tr>
      <tr><td>Parent Contact</td><td>{s.get('parent_contact','-')}</td></tr>
      <tr><td>Aadhaar No.</td><td>{s.get('aadhaar_number','-')}</td></tr>
      <tr><td>Bank Account</td><td>{s.get('bank_account','-')}</td></tr>
      <tr><td>Bank Name</td><td>{s.get('bank_name','-')}</td></tr>
      <tr><td>IFSC Code</td><td>{s.get('ifsc_code','-')}</td></tr>
    </table>
    <hr class="divider">

    <div class="sec-title">Attendance</div>
    <div class="att-row">
      <span>Total Days: <b>{total_d}</b></span>
      <span>Present: <b>{pres_d}</b></span>
      <span>Absent: <b>{total_d - pres_d}</b></span>
      <span>Percentage: <b>{att_pct}%</b></span>
    </div>
    <hr class="divider">

    <div class="sec-title">Marks Summary</div>
    <table class="mk-tbl">
      <thead>
        <tr>
          <th class="tl">Subject</th>
          <th>Semester</th>
          <th>Written</th>
          <th>Participation</th>
          <th>Total</th>
          <th>Grade</th>
        </tr>
      </thead>
      <tbody>
        {marks_rows_p1 if marks_rows_p1 else '<tr><td colspan="6" style="text-align:center;color:#94a3b8;padding:10px;">No marks entered yet.</td></tr>'}
      </tbody>
    </table>
    {docs_section}
  </div>
  <div class="p1-footer">
    <span>Rajivnagar Primary School, Rajivnagar, Kadi, Mehsana, Gujarat</span>
    <span>Generated: {str(date.today())}</span>
  </div>
</div>


<!-- ══════════════ PAGE 2: Gujarati Patrak-F ══════════════ -->
<div class="page p2">
  <div class="p2-inner">
    <div class="p2-title">શાળાકીય સર્વગ્રાહી મૂલ્યાંકન</div>
    <div class="p2-sub">જિલ્લો શિક્ષણ સમિતિ મહેસાણા</div>
    <span class="p2-dotline"></span>
    <div class="p2-main-title">પ્રગતિ પત્રક</div>
    <div class="p2-f">પત્રક-F</div>
    <div class="p2-progress">(PROGRESS REPORT)</div>

    <div class="p2-row">
      <span class="p2-lbl">ધોરણ :</span>
      <span class="p2-val" style="max-width:110px;">{cls_num}</span>
      <span style="min-width:28px;"></span>
      <span class="p2-lbl" style="min-width:70px;">વર્ગ :</span>
      <span class="p2-val">{s.get('section','-')}</span>
    </div>

    <div style="text-align:center;font-size:12px;margin:8px 0 10px;">
      સને : 20 / 20
    </div>

    <div class="p2-row">
      <span class="p2-lbl">શાળાનું નામ :</span>
      <span class="p2-val"><b>રાજીવનગર પ્રાથમિક શાળા</b></span>
    </div>

    <div class="p2-row">
      <span class="p2-lbl">તાલુકો :</span>
      <span class="p2-val" style="max-width:130px;"><b>કડી</b></span>
      <span style="min-width:28px;"></span>
      <span class="p2-lbl" style="min-width:70px;">જિલ્લો :</span>
      <span class="p2-val"><b>મહેસાણા</b></span>
    </div>

    <div style="font-size:12px;margin:8px 0 4px;">નામ :</div>
    <div class="p2-name-row">
      <div class="p2-name-col">
        <span class="p2-nval">{s.get('surname','')}</span>
        <div class="p2-nlbl">(અટક)</div>
      </div>
      <div class="p2-name-col">
        <span class="p2-nval">{s.get('name','')}</span>
        <div class="p2-nlbl">(વિદ્યાર્થીનું નામ)</div>
      </div>
      <div class="p2-name-col">
        <span class="p2-nval">{s.get('father_name','')}</span>
        <div class="p2-nlbl">(પિતાનું નામ)</div>
      </div>
    </div>

    <div class="p2-row" style="margin-top:10px;">
      <span class="p2-lbl">સરનામું :</span>
      <span class="p2-val"><b>{s.get('address','')}</b></span>
    </div>
    <div class="p2-row">
      <span class="p2-lbl">પિતા / વાલીનું નામ :</span>
      <span class="p2-val"><b>{s.get('father_name','')}</b></span>
    </div>
    <div class="p2-row">
      <span class="p2-lbl">માતાનું નામ :</span>
      <span class="p2-val"><b>{s.get('mother_name','')}</b></span>
    </div>
    <div class="p2-row">
      <span class="p2-lbl">જ.ર.નંબર :</span>
      <span class="p2-val" style="max-width:130px;"><b>{s.get('gr_number','-')}</b></span>
      <span style="min-width:28px;"></span>
      <span class="p2-lbl" style="min-width:100px;">જન્મ તારીખ :</span>
      <span class="p2-val"><b>{dob_fmt}</b></span>
    </div>
    <div class="p2-row">
      <span class="p2-lbl">સંપર્ક નંબર :</span>
      <span class="p2-val"><b>{s.get('parent_contact','')}</b></span>
    </div>

    <hr class="p2-divider">
    <div class="p2-bank">
      <div>ખાતા નંબર :- {s.get('bank_account','')}</div>
      <div>બેંકનું નામ :- {s.get('bank_name','')}</div>
      <div>IFSC Code:- {s.get('ifsc_code','')}</div>
    </div>

    <div class="p2-spacer"></div>

    <div class="p2-sigs" style="margin-bottom: 10%">
      <div class="p2-sig">
        <div class="p2-sigline"></div>
        વર્ગશિક્ષકની સહી
      </div>
      <div class="p2-sig">
        <div class="p2-sigline"></div>
        આચાર્યની સહી
      </div>
    </div>
  </div>
</div>


<!-- ══════════════ PAGE 3: Gujarati Marks Table ══════════════ -->
<div class="page p3">
  <div class="p3-inner">
    <div class="p3-title">શાળાકીય સર્વગ્રાહી મૂલ્યાંકન</div>
    <div class="p3-info">
      વિદ્યાર્થીનું નામ:&nbsp;
      <span class="p3-info-val">{s.get('surname','')} {s.get('name','')} {s.get('father_name','')}</span>
      &nbsp;&nbsp; ધોરણ:&nbsp;<span class="p3-info-val" style="min-width:28px">{cls_num}</span>
      &nbsp;&nbsp; વર્ગ:&nbsp;<span class="p3-info-val" style="min-width:28px">{s.get('section','-')}</span>
      &nbsp;&nbsp; હાજરીપત્રક નંબર:&nbsp;<span class="p3-info-val" style="min-width:36px">{s.get('attendance_register_no','-')}</span>
      &nbsp;&nbsp; કુલ હાજર દિવસ:&nbsp;<span class="p3-info-val" style="min-width:36px">{pres_d}</span>
    </div>

    <table class="guj-tbl">
      <thead>
        <tr>
          <th style="width:38px">ક્રમ</th>
          <th class="tl">વિષય</th>
          <th>કુલ<br>ગુણ</th>
          <th>મેળવેલ<br>ગુણ</th>
          <th>મેળવેલ<br>ગ્રેડ</th>
          <th>વિષયના સંદર્ભમાં<br>વિશેષ નોંધ</th>
        </tr>
      </thead>
      <tbody>
        {marks_rows_p3 if marks_rows_p3 else '<tr><td colspan="6" style="text-align:center;padding:10px;">ગુણ ભર્યા નથી.</td></tr>'}
        <tr class="tot-row">
          <td colspan="2">કુલ ગુણ</td>
          <td>{grand_max}</td>
          <td>{grand_total}</td>
          <td>{overall_grade}</td>
          <td></td>
        </tr>
        <tr>
          <td>{row_count}</td>
          <td class="tl">મેળવેલ એકંદર ગ્રેડ</td>
          <td></td>
          <td>{overall_pct}%</td>
          <td>{overall_grade}</td>
          <td></td>
        </tr>
      </tbody>
    </table>

    <div class="p3-spacer"></div>

    <div class="p3-bottom">
      <span>સ્થળ :&nbsp;<span class="p3-bottom-val"></span></span>
      <span style="margin-left: 30%">વાલીની સહી :&nbsp;<span class="p3-bottom-val"></span></span>
    </div>
    <div style="font-size:15px;margin-top:5px; margin-bottom:65%;">
      <span>તારીખ :&nbsp;<span class="p3-bottom-val"></span></span>
    </div>

    <div class="p3-sigs" style="margin-bottom:10%;">
      <div class="p3-sig">
        <div class="p3-sigline"></div>
        વર્ગશિક્ષકની સહી
      </div>
      <div class="p3-sig">
        <div class="p3-sigline"></div>
        આચાર્યની સહી
      </div>
    </div>
  </div>
</div>

</body>
</html>"""

    pdf_bytes = HTML(string=html).write_pdf()
    buf = io.BytesIO(pdf_bytes)
    buf.seek(0)
    fname = (s.get('name') or 'Student').replace(' ', '_')
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'Student_{fname}.pdf')


# ─────────────────────────────────────────────
# EXCEL GENERATION
# ─────────────────────────────────────────────

def _generate_excel(my_class, semester):
    students  = (supabase.table('students').select('*').eq('class', my_class).order('roll_no').execute().data) or []
    subjects  = (supabase.table('subjects').select('*').eq('class', my_class).order('created_at').execute().data) or []
    all_marks = (supabase.table('marks').select('*').eq('semester', semester).execute().data) or []
    ml = {f"{m['student_id']}_{m['subject_id']}": m for m in all_marks}

    wb    = Workbook()
    thin  = Side(style='thin')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill('solid', fgColor='1a4f8a')
    hfont = Font(color='FFFFFF', bold=True, size=11)
    bfont = Font(bold=True, size=11)
    nfont = Font(size=11)

    ws = wb.active
    ws.title = 'Marks'
    ws['A1'] = 'RAJIVNAGAR PRIMARY SCHOOL'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    ws['A2'] = f'Class: {my_class}    Semester: {semester}    Year: 2025-26'
    ws['A2'].font = bfont
    ws.merge_cells('A2:C2')

    hdrs = ['Sr.', 'Roll', 'Student Name'] + [s['subject_name'] for s in subjects] + ['Total']
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.border = brd
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 14 if ci > 3 else ([5, 8, 28][ci-1])

    for ri, s in enumerate(students):
        row = 4 + ri; tot = 0
        vals = [ri+1, s['roll_no'], s['name']]
        for subj in subjects:
            key = f"{s['id']}_{subj['id']}"
            if key in ml:
                v = ml[key]['written_marks'] + ml[key]['participation_marks']
                vals.append(v); tot += v
            else: vals.append('-')
        vals.append(tot)
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = bfont if ci == len(vals) else nfont; cell.border = brd
            cell.alignment = Alignment(horizontal='center' if ci != 3 else 'left')

    ws2 = wb.create_sheet('Attendance')
    ws2['A1'] = f'RAJIVNAGAR PRIMARY SCHOOL — Attendance — {my_class}'
    ws2['A1'].font = Font(bold=True, size=13)
    ws2.merge_cells('A1:G1')
    for ci, h in enumerate(['Sr.', 'Roll', 'Name', 'Total Days', 'Present', 'Absent', 'Percentage'], 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.border = brd
        cell.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[cell.column_letter].width = [5, 8, 28, 12, 10, 10, 14][ci-1]

    all_att = (supabase.table('attendance').select('*').execute().data) or []
    for ri, s in enumerate(students):
        s_att = [a for a in all_att if a['student_id'] == s['id']]
        td = len(s_att); pr = len([a for a in s_att if a['status'] == 'P']); ab = td - pr
        pct = f"{round(pr/td*100)}%" if td > 0 else '-'
        for ci, v in enumerate([ri+1, s['roll_no'], s['name'], td, pr, ab, pct], 1):
            cell = ws2.cell(row=3+ri, column=ci, value=v)
            cell.font = nfont; cell.border = brd
            cell.alignment = Alignment(horizontal='center' if ci != 3 else 'left')

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'Rajivnagar_{my_class}_Sem{semester}.xlsx')


def _build_marks_lookup(my_class, semester):
    students  = (supabase.table('students').select('*').eq('class', my_class).order('roll_no').execute().data) or []
    subjects  = (supabase.table('subjects').select('*').eq('class', my_class).order('created_at').execute().data) or []
    all_marks = (supabase.table('marks').select('*').eq('semester', semester).execute().data) or []
    marks_data = {}
    for m in all_marks:
        sid = m['student_id']; subj_id = m['subject_id']
        if sid not in marks_data: marks_data[sid] = {}
        marks_data[sid][subj_id] = {'written': m['written_marks'], 'participation': m['participation_marks']}
    return students, subjects, marks_data


def _generate_gunslip_excel(my_class, semester):
    students, subjects, marks_data = _build_marks_lookup(my_class, semester)
    output = generate_gun_slip('રાજીવનગર પ્રાથમિક શાળા', 'કડી', my_class, semester, '2025-26', students, subjects, marks_data)
    cls_num = my_class.replace('Class ', '')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'GUN_SLIP_STD_{cls_num}_Sem{semester}.xlsx')


def _generate_parinam_excel(my_class, semester):
    students, subjects, marks_data = _build_marks_lookup(my_class, semester)
    all_att = (supabase.table('attendance').select('*').execute().data) or []
    output  = generate_parinam('રાજીવનગર પ્રાથમિક શાળા', 'કડી', my_class, semester, '2025-26', students, subjects, marks_data, all_att)
    cls_num = my_class.replace('Class ', '')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'PARINAM_STD_{cls_num}_Sem{semester}.xlsx')


# ─────────────────────────────────────────────
# TEACHER ROUTES
# ─────────────────────────────────────────────

@app.route('/teacher/dashboard')
def teacher_dashboard():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    my_class = session['class']; today = str(date.today())
    students = get_students_with_att(my_class); total = len(students)
    ids = [s['id'] for s in students]
    att_today = (supabase.table('attendance').select('*').eq('date', today).execute().data) or []
    my_att  = [a for a in att_today if a['student_id'] in ids]
    present = len([a for a in my_att if a['status'] == 'P'])
    absent  = len([a for a in my_att if a['status'] == 'A'])
    low_list = [{'name': s['name'], 'roll': s['roll_no'], 'pct': s['att_pct']} for s in students if s['att_pct'] < 75]
    return render_template('teacher/dashboard.html', total=total, present=present, absent=absent,
                           low_count=len(low_list), low_list=low_list[:5], today=today)


@app.route('/teacher/attendance')
def teacher_attendance():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    my_class = session['class']; today = str(date.today())
    students = get_students_with_att(my_class)
    att_today = (supabase.table('attendance').select('*').eq('date', today).execute().data) or []
    ids     = [s['id'] for s in students]
    att_map = {a['student_id']: a['status'] for a in att_today if a['student_id'] in ids}
    return render_template('teacher/attendance.html', students=students, att_map=att_map, today=today)


@app.route('/teacher/attendance/save', methods=['POST'])
def save_attendance():
    if not logged_in() or not is_teacher(): return jsonify({'error': 'Login required.'}), 401
    _save_attendance_data(request.json or {}); return jsonify({'success': True})


@app.route('/teacher/students')
def teacher_students():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    return render_template('teacher/students.html', students=get_students_with_att(session['class']))


@app.route('/teacher/students/add', methods=['POST'])
def add_student():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    return _add_student_logic(session['class'], 'teacher_students')


@app.route('/teacher/students/edit', methods=['POST'])
def edit_student():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    return _edit_student_logic('teacher_students')


@app.route('/teacher/students/delete', methods=['POST'])
def delete_student():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    supabase.table('students').delete().eq('id', request.form.get('student_id')).execute()
    return redirect(url_for('teacher_students'))


@app.route('/teacher/student/detail/<student_id>')
def student_detail(student_id):
    if not logged_in(): return jsonify({'error': 'Login required'}), 401
    s = get_student_full(student_id)
    return jsonify(s) if s else (jsonify({'error': 'Not found'}), 404)


@app.route('/teacher/subjects')
def teacher_subjects():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    subjects = (supabase.table('subjects').select('*').eq('class', session['class']).order('created_at').execute().data) or []
    return render_template('teacher/subjects.html', subjects=subjects)


@app.route('/teacher/subjects/add', methods=['POST'])
def add_subject():
    if not logged_in() or not is_teacher():
        return redirect(url_for('login'))

    name = request.form.get('subject_name', '').strip()
    max_w = int(request.form.get('max_written', 40) or 40)
    max_p = int(request.form.get('max_participation', 20) or 20)
    semester = request.form.get('semester', '1')

    if name:
        existing = supabase.table('subjects')\
            .select('id')\
            .eq('class', session['class'])\
            .eq('subject_name', name)\
            .eq('semester', semester)\
            .execute()

        if not existing.data:
            supabase.table('subjects').insert({
                'class': session['class'],
                'subject_name': name,
                'teacher_id': session['user_id'],
                'max_written': max_w,
                'max_participation': max_p,
                'semester': semester
            }).execute()

    return redirect(url_for('teacher_subjects'))


@app.route('/teacher/subjects/delete', methods=['POST'])
def delete_subject():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    supabase.table('subjects').delete().eq('id', request.form.get('subject_id')).execute()
    return redirect(url_for('teacher_subjects'))


@app.route('/teacher/marks')
def teacher_marks():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    my_class   = session['class']
    subject_id = request.args.get('subject_id', ''); semester = request.args.get('semester', '1')
    students   = (supabase.table('students').select('*').eq('class', my_class).order('roll_no').execute().data) or []
    subjects = (supabase.table('subjects')
    .select('*')
    .eq('class', my_class)
    .eq('semester', semester)
    .order('created_at')
    .execute().data) or []
    if not subject_id and subjects: subject_id = subjects[0]['id']
    current_subject = next((s for s in subjects if s['id'] == subject_id), None)
    marks_map = {}
    if subject_id:
        m_res = supabase.table('marks').select('*').eq('subject_id', subject_id).eq('semester', semester).execute()
        marks_map = {m['student_id']: m for m in (m_res.data or [])}
    return render_template('teacher/marks.html', students=students, subjects=subjects,
                           current_subject=current_subject, subject_id=subject_id,
                           semester=semester, marks_map=marks_map)


@app.route('/teacher/marks/save', methods=['POST'])
def save_marks():
    if not logged_in() or not is_teacher():
        return jsonify({'error': 'Login required.'}), 401

    data = request.json or {}

    # 🔥 UPDATE SUBJECT MAX MARKS
    supabase.table('subjects').update({
        'max_written': int(data.get('max_written', 40)),
        'max_participation': int(data.get('max_participation', 20))
    }).eq('id', data.get('subject_id')).execute()

    # existing marks save
    _save_marks_data(data)

    return jsonify({'success': True})


@app.route('/teacher/download')
def teacher_download():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    students = (supabase.table('students').select('*').eq('class', session['class']).order('roll_no').execute().data) or []
    return render_template('teacher/download.html', students=students)


@app.route('/teacher/download/excel')
def download_excel():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    return _generate_excel(session['class'], request.args.get('semester', '1'))


@app.route('/teacher/download/pdf/<student_id>')
def download_student_pdf(student_id):
    if not logged_in(): return redirect(url_for('login'))
    return _generate_student_pdf(student_id)


@app.route('/teacher/download/gunslip')
def download_gunslip():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    return _generate_gunslip_excel(session['class'], request.args.get('semester', '1'))


@app.route('/teacher/download/parinam')
def download_parinam():
    if not logged_in() or not is_teacher(): return redirect(url_for('login'))
    return _generate_parinam_excel(session['class'], request.args.get('semester', '1'))

from datetime import date, timedelta
import io
from flask import send_file, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

@app.route('/teacher/download/attendance-30days')
def teacher_download_attendance_30days():
    if not logged_in() or not is_teacher():
        return redirect(url_for('login'))

    my_class = session['class']

    # 🔹 Students
    students = supabase.table('students')\
        .select('*')\
        .eq('class', my_class)\
        .order('roll_no')\
        .execute().data or []

    # 🔹 Date range
    end_date = date.today()
    start_date = end_date - timedelta(days=29)

    # 🔹 Attendance
    all_att = supabase.table('attendance')\
        .select('*')\
        .gte('date', str(start_date))\
        .lte('date', str(end_date))\
        .execute().data or []

    att_map = {(a['student_id'], a['date']): a['status'] for a in all_att}

    # 🔹 Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Register"

    # Styles
    header_font = Font(bold=True, size=14)
    sub_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 🔹 Header
    start_str = start_date.strftime('%d-%m-%Y')
    end_str = end_date.strftime('%d-%m-%Y')

    ws['A1'] = 'RAJIVNAGAR PRIMARY SCHOOL'
    ws['A2'] = f'Class: {my_class}'
    ws['A3'] = 'Attendance Register (Last 30 Days)'
    ws['A4'] = f'From: {start_str} To: {end_str}'

    for r in range(1, 5):
        cell = ws[f'A{r}']
        cell.font = header_font if r == 1 else sub_font
        cell.alignment = center

    # 🔹 Dates
    dates = [start_date + timedelta(days=i) for i in range(30)]

    headers = ['Roll', 'Name'] + \
              [d.strftime('%d %b') for d in dates] + \
              ['Total', 'Present', 'Absent', '%']

    total_columns = len(headers)

    # Merge header
    for i in range(1, 5):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=total_columns)

    ws.append([])
    ws.append(headers)

    header_row = ws.max_row

    # Header styling
    for col in range(1, total_columns + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = Font(bold=True)
        c.alignment = center
        c.border = border

    # 🔹 Data fill
    for s in students:
        row = [s.get('roll_no'), s.get('name')]

        present = 0
        total = 0

        for d in dates:
            d_str = str(d)

            # Sunday skip
            if d.weekday() == 6:
                row.append('-')
                continue

            status = att_map.get((s['id'], d_str), '')

            if status == 'P':
                present += 1
                total += 1
                row.append('P')
            elif status == 'A':
                total += 1
                row.append('A')
            else:
                row.append('')

        absent = total - present
        pct = round((present / total) * 100, 1) if total > 0 else 0

        row += [total, present, absent, f"{pct}%"]
        ws.append(row)

    # 🔹 Styling data
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(1, total_columns + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border

            if cell.value == 'P':
                cell.fill = green_fill
            elif cell.value == 'A':
                cell.fill = red_fill

    # 🔹 Column width
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25

    for col in range(3, total_columns + 1):
        ws.column_dimensions[get_column_letter(col)].width = 6

    # 🔹 Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{my_class}_Attendance_Register_30Days.xlsx'
    )



# ─────────────────────────────────────────────
# PRINCIPAL ROUTES
# ────────────────────────────────────────────

@app.route('/principal/dashboard')
def principal_dashboard():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    today     = str(date.today())
    all_stu   = (supabase.table('students').select('*').execute().data) or []
    all_tea   = (supabase.table('users').select('*').eq('role', 'teacher').execute().data) or []
    att_today = (supabase.table('attendance').select('*').eq('date', today).execute().data) or []
    total_present = len([a for a in att_today if a['status'] == 'P'])
    class_summary = []
    for num in range(1, 9):
        cls     = f'Class {num}'
        cls_stu = [s for s in all_stu if s['class'] == cls]
        if not cls_stu: continue
        ids    = [s['id'] for s in cls_stu]
        pres   = len([a for a in att_today if a['student_id'] in ids and a['status'] == 'P'])
        marked = any(a['student_id'] in ids for a in att_today)
        teacher = next((t for t in all_tea if t.get('class_assigned') == cls), None)
        class_summary.append({'class': cls, 'teacher': teacher['name'] if teacher else 'Not Assigned',
                               'total': len(cls_stu), 'present': pres, 'att_marked': marked})
    return render_template('principal/dashboard.html', total_students=len(all_stu),
                           total_teachers=len(all_tea), total_present=total_present,
                           class_summary=class_summary, today=today)


@app.route('/principal/teachers')
def principal_teachers():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    teachers = (supabase.table('users').select('*').eq('role', 'teacher').execute().data) or []
    return render_template('principal/teachers.html', teachers=teachers)


@app.route('/principal/teachers/add', methods=['POST'])
def add_teacher():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    name     = request.form.get('name', '').strip()
    cls      = request.form.get('class_assigned', '').strip()
    if all([username, password, name]):
        if not supabase.table('users').select('id').eq('username', username).execute().data:
            supabase.table('users').insert({'name': name, 'username': username,
                'password_hash': generate_password_hash(password),
                'role': 'teacher', 'class_assigned': cls or None}).execute()
    return redirect(url_for('principal_teachers'))


@app.route('/principal/teachers/reset', methods=['POST'])
def reset_teacher_password():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    tid = request.form.get('teacher_id'); new = request.form.get('new_password', '').strip()
    if tid and new:
        supabase.table('users').update({'password_hash': generate_password_hash(new)}).eq('id', tid).execute()
    return redirect(url_for('principal_teachers'))


@app.route('/principal/teachers/delete', methods=['POST'])
def delete_teacher():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    tid = request.form.get('teacher_id')
    if tid: supabase.table('users').delete().eq('id', tid).execute()
    return redirect(url_for('principal_teachers'))


@app.route('/principal/students')
def principal_students():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    class_filter = request.args.get('class', '')
    students = get_students_with_att(class_filter if class_filter else None)
    return render_template('principal/students.html', students=students, class_filter=class_filter)


@app.route('/principal/students/add', methods=['POST'])
def principal_add_student():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    return _add_student_logic(request.form.get('class', '').strip(), 'principal_students')


@app.route('/principal/students/edit', methods=['POST'])
def principal_edit_student():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    return _edit_student_logic('principal_students')


@app.route('/principal/students/delete', methods=['POST'])
def principal_delete_student():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    supabase.table('students').delete().eq('id', request.form.get('student_id')).execute()
    return redirect(url_for('principal_students'))


@app.route('/principal/student/detail/<student_id>')
def principal_student_detail(student_id):
    if not logged_in() or not is_principal(): return jsonify({'error': 'Login required'}), 401
    s = get_student_full(student_id)
    return jsonify(s) if s else (jsonify({'error': 'Not found'}), 404)


@app.route('/principal/attendance')
def principal_attendance():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    class_filter = request.args.get('class', 'Class 1'); today = str(date.today())
    students  = get_students_with_att(class_filter)
    att_today = (supabase.table('attendance').select('*').eq('date', today).execute().data) or []
    ids     = [s['id'] for s in students]
    att_map = {a['student_id']: a['status'] for a in att_today if a['student_id'] in ids}
    classes = [f'Class {i}' for i in range(1, 9)]
    return render_template('principal/attendance.html', students=students, att_map=att_map,
                           today=today, class_filter=class_filter, classes=classes)


@app.route('/principal/attendance/save', methods=['POST'])
def principal_save_attendance():
    if not logged_in() or not is_principal(): return jsonify({'error': 'Login required'}), 401
    _save_attendance_data(request.json or {}); return jsonify({'success': True})


@app.route('/principal/marks')
def principal_marks():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    class_filter = request.args.get('class', 'Class 1')
    subject_id   = request.args.get('subject_id', ''); semester = request.args.get('semester', '1')
    students = (supabase.table('students').select('*').eq('class', class_filter).order('roll_no').execute().data) or []
    subjects = (supabase.table('subjects').select('*').eq('class', class_filter).order('created_at').execute().data) or []
    if not subject_id and subjects: subject_id = subjects[0]['id']
    current_subject = next((s for s in subjects if s['id'] == subject_id), None)
    marks_map = {}
    if subject_id:
        m_res = supabase.table('marks').select('*').eq('subject_id', subject_id).eq('semester', semester).execute()
        marks_map = {m['student_id']: m for m in (m_res.data or [])}
    classes = [f'Class {i}' for i in range(1, 9)]
    return render_template('principal/marks.html', students=students, subjects=subjects,
                           current_subject=current_subject, subject_id=subject_id, semester=semester,
                           marks_map=marks_map, class_filter=class_filter, classes=classes)


@app.route('/principal/download')
def principal_download():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    class_filter = request.args.get('class', 'Class 1')
    all_students = (supabase.table('students').select('*').order('class').order('roll_no').execute().data) or []
    classes = [f'Class {i}' for i in range(1, 9)]
    return render_template('principal/download.html', students=all_students,
                           class_filter=class_filter, classes=classes)


@app.route('/principal/download/excel')
def principal_download_excel():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    return _generate_excel(request.args.get('class', 'Class 1'), request.args.get('semester', '1'))


@app.route('/principal/download/pdf/<student_id>')
def principal_download_pdf(student_id):
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    return _generate_student_pdf(student_id)


@app.route('/principal/download/gunslip')
def principal_download_gunslip():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    return _generate_gunslip_excel(request.args.get('class', 'Class 1'), request.args.get('semester', '1'))


@app.route('/principal/download/parinam')
def principal_download_parinam():
    if not logged_in() or not is_principal(): return redirect(url_for('login'))
    return _generate_parinam_excel(request.args.get('class', 'Class 1'), request.args.get('semester', '1'))


from datetime import date, timedelta
import io
from flask import send_file, redirect, url_for, request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

@app.route('/principal/download/attendance-30days')
def principal_download_attendance_30days():
    if not logged_in() or not is_principal():
        return redirect(url_for('login'))

    class_name = request.args.get('class', 'Class 1').strip()

    # 🔹 Students
    students = supabase.table('students')\
        .select('*')\
        .ilike('class', class_name)\
        .order('roll_no')\
        .execute().data or []

    # 🔹 Date range
    end_date = date.today()
    start_date = end_date - timedelta(days=29)

    # 🔹 Attendance
    all_att = supabase.table('attendance')\
        .select('*')\
        .gte('date', str(start_date))\
        .lte('date', str(end_date))\
        .execute().data or []

    att_map = {(a['student_id'], a['date']): a['status'] for a in all_att}

    # 🔹 Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Register"

    # Styles
    header_font = Font(bold=True, size=14)
    sub_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 🔹 Header
    start_str = start_date.strftime('%d-%m-%Y')
    end_str = end_date.strftime('%d-%m-%Y')

    ws['A1'] = 'RAJIVNAGAR PRIMARY SCHOOL'
    ws['A2'] = f'Class: {class_name}'
    ws['A3'] = 'Attendance Register (Last 30 Days)'
    ws['A4'] = f'From: {start_str} To: {end_str}'

    for r in range(1, 5):
        cell = ws[f'A{r}']
        cell.font = header_font if r == 1 else sub_font
        cell.alignment = center

    # 🔹 Dates
    dates = [start_date + timedelta(days=i) for i in range(30)]

    headers = ['Roll', 'Name'] + \
              [d.strftime('%d %b') for d in dates] + \
              ['Total', 'Present', 'Absent', '%']

    total_columns = len(headers)

    # Merge header
    for i in range(1, 5):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=total_columns)

    ws.append([])
    ws.append(headers)

    header_row = ws.max_row

    # Style header row
    for col in range(1, total_columns + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = Font(bold=True)
        c.alignment = center
        c.border = border

    # 🔹 Fill Data
    for s in students:
        row = [s.get('roll_no'), s.get('name')]

        present = 0
        total = 0

        for d in dates:
            d_str = str(d)

            # Sunday skip
            if d.weekday() == 6:
                row.append('-')
                continue

            status = att_map.get((s['id'], d_str), '')

            if status == 'P':
                present += 1
                total += 1
                row.append('P')
            elif status == 'A':
                total += 1
                row.append('A')
            else:
                row.append('')

        absent = total - present
        pct = round((present / total) * 100, 1) if total > 0 else 0

        row += [total, present, absent, f"{pct}%"]
        ws.append(row)

    # 🔹 Apply styling to data
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(1, total_columns + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border

            if cell.value == 'P':
                cell.fill = green_fill
            elif cell.value == 'A':
                cell.fill = red_fill

    # 🔹 Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25

    for col in range(3, total_columns + 1):
        ws.column_dimensions[get_column_letter(col)].width = 6

    # 🔹 Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{class_name}_Attendance_Register_30Days.xlsx'
    )


if __name__ == '__main__':
    app.run(debug=True)
