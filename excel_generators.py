"""
Excel generators for Rajivnagar Shala Dashboard
GUN_SLIP format and PARINAM/MARKSHEET format
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─── Common helpers ──────────────────────────────────────────────

def thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def top_bottom_border():
    t = Side(style='thin')
    return Border(top=t, bottom=t)

def bold(size=10, color='000000'):
    return Font(name='Shruti', bold=True, size=size, color=color)

def normal(size=10, color='000000'):
    return Font(name='Shruti', size=size, color=color)

def center_align(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

def set_cell(ws, row, col, value, font=None, align=None, border=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font      = font
    if align:  cell.alignment = align
    if border: cell.border    = border
    if fill:   cell.fill      = fill
    return cell

def grey_fill():
    return PatternFill('solid', fgColor='D9D9D9')

def yellow_fill():
    return PatternFill('solid', fgColor='FFFF00')

def light_blue_fill():
    return PatternFill('solid', fgColor='BDD7EE')

def col_letter_to_num(letter):
    num = 0
    for c in letter.upper():
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num

def get_grade_local(total, out_of=60):
    if out_of == 0: return '-'
    p = (total / out_of) * 100
    if p >= 90:  return 'A+'
    elif p >= 75: return 'A'
    elif p >= 60: return 'B'
    elif p >= 45: return 'C'
    else:         return 'D'


# ─── Subject name helper ─────────────────────────────────────────
# FIX: Subject names are already in Gujarati in DB.
# Only map if it's an English name, otherwise return as-is (full name, no truncation)

GUJ_NAMES = {
    'Gujarati':           'ગુજ.',
    'English':            'અંગ.',
    'Hindi':              'હિ.',
    'Sanskrit':           'સં.',
    'Mathematics':        'ગણ.',
    'Science':            'વિ.',
    'Social Science':     'સા.વિ.',
    'Computer':           'કોમ.',
    'Art':                'ચિ.',
    'Physical Education': 'શ.શિ.',
}

def guj_short(name):
    """Return short Gujarati name. If name is already Gujarati, return as-is."""
    return GUJ_NAMES.get(name, name)  # FIX: return full name if not in dict


def safe_sheet_name(name, max_len=28):
    """Safe Excel sheet name — no truncation that breaks Gujarati."""
    invalid = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in invalid:
        name = name.replace(ch, '')
    return name[:max_len]


# ─── GUN_SLIP Generator ──────────────────────────────────────────

def generate_gun_slip(school_name, taluko, std_class, semester, year,
                      students, subjects, marks_data):
    wb = Workbook()
    tb = thin_border()
    exam_name = 'પ્રથમ સત્રાંત ૫રીક્ષા' if semester == '1' else 'દ્વિતીય સત્રાંત ૫રીક્ષા'

    cls_num        = std_class.replace('Class ', '')
    total_students = len(students)

    # ── SHEET: DATA ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'DATA'
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 8

    info = [
        (2, 'B', 'C', 'સમિતિ :',           'જિલ્લા શિક્ષણ સમિતિ, મહેસાણા'),
        (3, 'B', 'C', 'શાળાનું પૂરું નામ :', school_name),
        (4, 'B', 'C', 'તાલુકો :',          taluko),
        (5, 'B', 'C', 'વર્ષ :',            year),
        (6, 'B', 'C', 'ધોરણ :',            cls_num),
        (7, 'B', 'C', 'રજીસ્ટર સંખ્યા :',  str(total_students)),
        (8, 'B', 'C', '૫રીક્ષા :',         exam_name),
    ]
    for row, lc, vc, label, val in info:
        set_cell(ws, row, col_letter_to_num(lc), label, font=bold())
        set_cell(ws, row, col_letter_to_num(vc), val,   font=normal())

    set_cell(ws, 9, 2, 'ક્રમ',              font=bold(), align=center_align(), border=tb)
    set_cell(ws, 9, 3, 'વિઘાર્થીનું નામ',  font=bold(), align=center_align(), border=tb)
    for i, s in enumerate(students):
        row = 10 + i
        set_cell(ws, row, 2, i + 1,    font=normal(), align=center_align(), border=tb)
        set_cell(ws, row, 3, s['name'], font=normal(), align=left_align(),   border=tb)

    # ── SHEETS: One per subject ──────────────────────────────────
    for subj in subjects:
        sname = subj['subject_name']
        gname = guj_short(sname)  # FIX: full name, no truncation
        # FIX: safe sheet name — no truncation that breaks Gujarati conjuncts
        ws_s  = wb.create_sheet(safe_sheet_name(sname))

        ws_s.column_dimensions['B'].width = 5
        ws_s.column_dimensions['C'].width = 22
        for col_idx in range(4, 20):
            ws_s.column_dimensions[get_column_letter(col_idx)].width = 5
        ws_s.column_dimensions['R'].width = 8
        ws_s.column_dimensions['S'].width = 8

        set_cell(ws_s, 3, 2, 'જિલ્લા શિક્ષણ સમિતિ, મહેસાણા', font=bold(11))
        set_cell(ws_s, 4, 2, f'{exam_name} : {year}                          લેખિત કસોટી  ગુણ ૫ત્રક', font=bold(10))
        set_cell(ws_s, 5, 2, f'શાળાનું નામ :  {school_name}    તા. :  {taluko}', font=normal())
        set_cell(ws_s, 6, 2, f'ધોરણ :  {cls_num}            રજીસ્ટર સંખ્યા : {total_students}', font=normal())
        set_cell(ws_s, 6, 16, 'હાજર સંખ્યા :', font=normal())

        set_cell(ws_s, 8, 2, 'ક્રમ',             font=bold(), align=center_align(), border=tb)
        set_cell(ws_s, 8, 3, 'વિઘાર્થીનું નામ',  font=bold(), align=center_align(), border=tb)
        set_cell(ws_s, 8, 4, f'વિષય :- {gname}', font=bold(), align=center_align(), border=tb)
        ws_s.merge_cells('D8:R8')

        set_cell(ws_s, 9, 2, '', border=tb)
        set_cell(ws_s, 9, 3, '', border=tb)
        for q in range(1, 15):
            set_cell(ws_s, 9, 3 + q, q, font=normal(), align=center_align(), border=tb)
        set_cell(ws_s, 9, 18, 'કુલ',       font=bold(),   align=center_align(), border=tb)
        set_cell(ws_s, 9, 19, '40\nમાંથી', font=normal(), align=center_align(wrap=True), border=tb)

        for i, s in enumerate(students):
            row = 10 + i
            sid = s['id']
            written = 0
            if sid in marks_data and subj['id'] in marks_data[sid]:
                written = marks_data[sid][subj['id']].get('written', 0)
            set_cell(ws_s, row, 2, i + 1,    font=normal(), align=center_align(), border=tb)
            set_cell(ws_s, row, 3, s['name'], font=normal(), align=left_align(),   border=tb)
            for q in range(1, 14):
                set_cell(ws_s, row, 3 + q, '', border=tb)
            set_cell(ws_s, row, 17, written if written else '', font=normal(), align=center_align(), border=tb)
            set_cell(ws_s, row, 18, written if written else '', font=bold(),   align=center_align(), border=tb)
            set_cell(ws_s, row, 19, 40,                         font=normal(), align=center_align(), border=tb)

        last_student_row = 10 + len(students)
        signature_row = last_student_row + 2
    
    # Left signature: પરીક્ષણકર્તા સહી (Examiner's Signature)
        set_cell(ws_s, signature_row, 2, 'પરીક્ષકની સહી', font=normal(), align=center_align())
        ws_s.merge_cells(f'B{signature_row}:H{signature_row}')
    
    # Right signature: સામીશીતો સહી (Principal's Signature)
        set_cell(ws_s, signature_row, 14, 'આચાર્યશ્રીની સહી', font=normal(), align=center_align())
        ws_s.merge_cells(f'N{signature_row}:U{signature_row}')

    # ── SHEET: TOTAL ─────────────���───────────────────────────────
    ws_t = wb.create_sheet('TOTAL')
    ws_t.column_dimensions['B'].width = 5
    ws_t.column_dimensions['C'].width = 22
    for ci in range(4, 20):
        ws_t.column_dimensions[get_column_letter(ci)].width = 8

    set_cell(ws_t, 3, 2, 'જિલ્લા શિક્ષણ સમિતિ, મહેસાણા', font=bold(11))
    set_cell(ws_t, 4, 2, f'{exam_name} : {year}                          લેખિત કસોટી  ગુણ ૫ત્રક', font=bold(10))
    set_cell(ws_t, 5, 2, f'શાળાનું નામ :  {school_name}    તા. :  {taluko}', font=normal())
    set_cell(ws_t, 6, 2, f'ધોરણ :  {cls_num}            રજીસ્ટર સંખ્યા : {total_students}', font=normal())
    set_cell(ws_t, 6, 12, 'હાજર સંખ્યા :', font=normal())

    set_cell(ws_t, 8, 2, 'ક્રમ',            font=bold(), align=center_align(), border=tb)
    set_cell(ws_t, 8, 3, 'વિઘાર્થીનું નામ', font=bold(), align=center_align(), border=tb)

    col = 4
    subj_cols = {}
    for subj in subjects:
        gname = guj_short(subj['subject_name'])  # FIX
        set_cell(ws_t, 8, col, gname, font=bold(), align=center_align(), border=tb)
        ws_t.merge_cells(f'{get_column_letter(col)}8:{get_column_letter(col+1)}8')
        set_cell(ws_t, 9, col,   80, font=normal(), align=center_align(), border=tb)
        set_cell(ws_t, 9, col+1, 40, font=normal(), align=center_align(), border=tb)
        subj_cols[subj['id']] = col
        col += 2

    for i, s in enumerate(students):
        row = 10 + i
        sid = s['id']
        set_cell(ws_t, row, 2, i + 1,    font=normal(), align=center_align(), border=tb)
        set_cell(ws_t, row, 3, s['name'], font=normal(), align=left_align(),   border=tb)
        for subj in subjects:
            c = subj_cols[subj['id']]
            written = 0
            if sid in marks_data and subj['id'] in marks_data[sid]:
                written = marks_data[sid][subj['id']].get('written', 0)
            set_cell(ws_t, row, c,   written if written else '', font=normal(), align=center_align(), border=tb)
            set_cell(ws_t, row, c+1, written if written else '', font=normal(), align=center_align(), border=tb)

    # ── SHEET: 20 GUN ─────────────────────────────────────────────
    ws_p = wb.create_sheet('20 GUN')
    ws_p.column_dimensions['B'].width = 5
    ws_p.column_dimensions['C'].width = 22
    for ci in range(4, 15):
        ws_p.column_dimensions[get_column_letter(ci)].width = 8

    set_cell(ws_p, 3, 2, 'જિલ્લા શિક્ષણ સમિતિ, મહેસાણા', font=bold(11))
    set_cell(ws_p, 4, 2, f'{exam_name} : {year}           ધોરણ :  {cls_num}', font=bold(10))
    set_cell(ws_p, 5, 2, f'શાળાનું નામ :  {school_name}    તા. :  {taluko}', font=normal())
    set_cell(ws_p, 6, 2, 'વર્ગખંડમાં વિઘાર્થીની સહભાગિતા આઘારે મૂલ્યાંકન ૫ત્રક', font=bold(10))

    set_cell(ws_p, 8, 2, 'ક્રમ',            font=bold(), align=center_align(), border=tb)
    set_cell(ws_p, 8, 3, 'વિઘાર્થીનું નામ', font=bold(), align=center_align(), border=tb)

    col = 4
    subj_part_cols = {}
    for subj in subjects:
        gname = guj_short(subj['subject_name'])  # FIX
        set_cell(ws_p, 8, col, gname, font=bold(), align=center_align(True), border=tb)
        set_cell(ws_p, 9, col, 20,    font=normal(), align=center_align(), border=tb)
        subj_part_cols[subj['id']] = col
        col += 1

    for i, s in enumerate(students):
        row = 10 + i
        sid = s['id']
        set_cell(ws_p, row, 2, i + 1,    font=normal(), align=center_align(), border=tb)
        set_cell(ws_p, row, 3, s['name'], font=normal(), align=left_align(),   border=tb)
        for subj in subjects:
            c = subj_part_cols[subj['id']]
            part = 0
            if sid in marks_data and subj['id'] in marks_data[sid]:
                part = marks_data[sid][subj['id']].get('participation', 0)
            set_cell(ws_p, row, c, part if part else '', font=normal(), align=center_align(), border=tb)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── PARINAM / MARKSHEET Generator ──────────────────────────────

def generate_parinam(school_name, taluko, std_class, semester, year,
                     students, subjects, marks_data, all_att):
    wb  = Workbook()
    tb  = thin_border()
    cls_num   = std_class.replace('Class ', '')
    exam_name = 'પ્રથમ સત્રાંત' if semester == '1' else 'દ્વિતીય સત્રાંત'

    def get_att(student_id):
        s_att = [a for a in all_att if a['student_id'] == student_id]
        total = len(s_att)
        pres  = len([a for a in s_att if a['status'] == 'P'])
        return total, pres

    # ── SHEET: DATA ───────────────────────────────────────────────
    ws_d = wb.active
    ws_d.title = 'DATA'

    for col, w in zip('BCDEFGHIJKLMNO', [22,8,30,14,12,10,10,18,18,18,22,22,30,16]):
        ws_d.column_dimensions[col].width = w

    total_students = len(students)

    info_rows = [
        (2,  [(3,'સમિતિ :',bold()),             (4,'જિલ્લા શિક્ષણ સમિતિ, મહેસાણા',normal()),  (7,'ગામ :-',bold())]),
        (3,  [(2,'શાળાનું પૂરું નામ :',bold()),  (4,school_name,normal()), (7,'ધોરણ :-',bold()), (9,cls_num,bold())]),
        (4,  [(2,'તાલુકો :',bold()),             (4,taluko,normal()),      (7,'વર્ગ :-',bold()),  (9,'',normal())]),
        (5,  [(2,'જીલ્લો :',bold()),             (4,'મહેસાણા',normal()),   (7,'વર્ષ :-',bold()),  (9,year,normal())]),
        (6,  [(2,'સી.આર.સી. :',bold()),          (4,'',normal()),          (7,'પરિણામ તારીખ :-',bold())]),
        (7,  [(2,'પે સેન્ટર :',bold()),          (4,'',normal()),          (7,'કાર્ય દિવસ :-',bold())]),
        (8,  [(2,'વર્ગ શિક્ષકનું નામ :',bold()), (4,'',normal()),          (7,'સત્ર :-',bold()),  (9,semester,bold())]),
        (9,  [(2,'વર્ગની કુલ રજીસ્ટર સંખ્યા :',bold()), (4,total_students,normal()), (7,'શાળાનો ડાયસ કોડ :-',bold())]),
    ]
    for row, cells in info_rows:
        for col, val, fnt in cells:
            set_cell(ws_d, row, col, val, font=fnt)

    hdrs = ['ક્રમ','વિદ્યાર્થીનું નામ','જી.આર. નંબર','જન્મ તારીખ',
            'હાજર દિવસ','કુમાર/કન્યા','જાતિ','આધાર ડાયસ','બેંક એકા.',
            'આધાર કાર્ડ','પિતા / વાલી','માતા','સરનામું','મોબાઇલ']
    for ci, h in enumerate(hdrs, 2):
        set_cell(ws_d, 10, ci, h, font=bold(9), align=center_align(True), border=tb)

    for i, s in enumerate(students):
        row = 11 + i
        total_d, pres_d = get_att(s['id'])
        vals = [
            i + 1, s.get('name',''), s.get('gr_number',''),
            s.get('dob',''), pres_d, s.get('gender',''),
            s.get('caste',''), s.get('aadhaar_number',''),
            s.get('bank_account',''), s.get('aadhaar_number',''),
            s.get('father_name',''), s.get('mother_name',''),
            s.get('address',''), s.get('parent_contact',''),
        ]
        for ci, v in enumerate(vals, 2):
            set_cell(ws_d, row, ci, v, font=normal(9), border=tb,
                     align=center_align() if ci != 3 else left_align())

    # ── SHEET: MARKSHEET ──────────────────────────────────────────
    ws_m = wb.create_sheet('MARKSHEET')

    for ci, w in zip(range(2, 11),  [5, 22, 5, 5, 5, 10, 14, 14, 14]):
        ws_m.column_dimensions[get_column_letter(ci)].width = w
    for ci, w in zip(range(12, 21), [3, 22, 5, 5, 5, 10, 14, 14, 14]):
        ws_m.column_dimensions[get_column_letter(ci)].width = w
    ws_m.column_dimensions['K'].width = 2

    def write_student_block(ws, s_data, base_col, base_row, subj_list):
        B = base_col
        r = base_row
        sid = s_data['id']
        total_d, pres_d = get_att(sid)
        dob_str = s_data.get('dob') or '   '

        # Header
        set_cell(ws, r,   B,   'જીલ્લા શિક્ષણ સમિતિ, મહેસાણા', font=bold(10))
        set_cell(ws, r+1, B,   f'{exam_name}',                    font=bold(10))
        set_cell(ws, r+1, B+4, 'પરીક્ષા પરિણામ સને :',           font=normal(9))
        set_cell(ws, r+1, B+7, year,                               font=bold(10))
        set_cell(ws, r+2, B,   'શાળાનું નામ :',                   font=normal(9))
        set_cell(ws, r+2, B+4, school_name,                        font=bold(10))
        set_cell(ws, r+3, B,   'રોલ નંબર :',                      font=normal(9))
        set_cell(ws, r+3, B+5, s_data.get('roll_no', ''),          font=bold(10))
        set_cell(ws, r+3, B+7, 'ધોરણ :',                          font=normal(9))
        set_cell(ws, r+3, B+8, cls_num,                            font=bold(10))
        set_cell(ws, r+4, B,   'વિઘાર્થીનું નામ :',               font=normal(9))
        full_name = f"{s_data.get('surname','')} {s_data.get('name','')} {s_data.get('father_name','')}".strip()
        set_cell(ws, r+4, B+3, full_name,                          font=bold(10))
        set_cell(ws, r+5, B,   'જન્મ તારીખ :',                    font=normal(9))
        set_cell(ws, r+5, B+4, dob_str,                            font=normal(9))
        set_cell(ws, r+5, B+7, 'જ.ર.નં.',                         font=normal(9))
        set_cell(ws, r+5, B+8, s_data.get('gr_number', ''),        font=normal(9))
        set_cell(ws, r+6, B,   'આધાર ડાયસ નંબર :',                font=normal(9))
        set_cell(ws, r+6, B+7, s_data.get('aadhaar_number', ''),   font=normal(9))
        set_cell(ws, r+7, B,   'બેંક એકા. નંબર :',                font=normal(9))
        set_cell(ws, r+7, B+7, s_data.get('bank_account', ''),     font=normal(9))
        set_cell(ws, r+8, B,   'આધાર કાર્ડ નંબર :',               font=normal(9))
        set_cell(ws, r+8, B+7, s_data.get('aadhaar_number', ''),   font=normal(9))
        set_cell(ws, r+9, B,   'હાજર દિવસ :',                     font=normal(9))
        set_cell(ws, r+9, B+4, pres_d,                             font=normal(9))
        set_cell(ws, r+9, B+5, 'માંથી',                            font=normal(9))
        set_cell(ws, r+9, B+7, total_d,                            font=normal(9))

        # Table header
        hr   = r + 10
        hdrs = ['ક્રમ', 'વિષય', '', '', '', 'કુલ ગુણ', 'મેળવેલ ગુણ', 'મેળવેલ ગ્રેડ', 'વિ.સં.ન.']
        for ci, h in enumerate(hdrs):
            cell = ws.cell(row=hr, column=B+ci, value=h)
            cell.font      = bold(9)
            cell.border    = tb
            cell.alignment = center_align(True)
        ws.merge_cells(f'{get_column_letter(B+1)}{hr}:{get_column_letter(B+4)}{hr}')

        # Subject rows — FIX: use subject_name directly (already Gujarati)
        grand_total_max = 0
        grand_total_got = 0
        dr = hr + 1

        for idx, subj in enumerate(subj_list):
            written = 0; part = 0
            if sid in marks_data and subj['id'] in marks_data[sid]:
                written = marks_data[sid][subj['id']].get('written', 0)
                part    = marks_data[sid][subj['id']].get('participation', 0)
            total_marks = written + part
            max_marks   = 60
            grand_total_max += max_marks
            grand_total_got += total_marks
            grade = get_grade_local(total_marks, max_marks)

            # FIX: use subj['subject_name'] directly — no guj_name() truncation
            cell_vals = [idx+1, subj['subject_name'], '', '', '', max_marks, total_marks, grade, '']
            for ci, v in enumerate(cell_vals):
                cell = ws.cell(row=dr+idx, column=B+ci, value=v)
                cell.font      = normal(9)
                cell.border    = tb
                cell.alignment = center_align() if ci != 1 else left_align()
            ws.merge_cells(f'{get_column_letter(B+1)}{dr+idx}:{get_column_letter(B+4)}{dr+idx}')

        # Total row
        tr = dr + len(subj_list)
        total_vals = ['', 'કુલ ગુણ', '', '', '', grand_total_max, grand_total_got, '', '']
        for ci, v in enumerate(total_vals):
            cell = ws.cell(row=tr, column=B+ci, value=v)
            cell.font      = bold(9)
            cell.border    = tb
            cell.alignment = center_align()
        ws.merge_cells(f'{get_column_letter(B+1)}{tr}:{get_column_letter(B+4)}{tr}')

        # Overall grade row
        overall_grade = get_grade_local(grand_total_got, grand_total_max)
        pct_val = round(grand_total_got / grand_total_max * 100, 1) if grand_total_max else 0
        gr = tr + 1
        set_cell(ws, gr, B,   'મેળવેલ એકંદરે ગ્રેડ', font=bold(9),   border=tb)
        set_cell(ws, gr, B+5, '',                      font=normal(9), border=tb)
        set_cell(ws, gr, B+6, f'{pct_val}%',           font=bold(9),   border=tb, align=center_align())
        set_cell(ws, gr, B+7, overall_grade,           font=bold(9),   border=tb, align=center_align())
        set_cell(ws, gr, B+8, '',                      font=normal(9), border=tb)
        ws.merge_cells(f'{get_column_letter(B)}{gr}:{get_column_letter(B+4)}{gr}')

        # Signatures
        sr = gr + 2
        set_cell(ws, sr, B,   'વર્ગ શિક્ષકની સહી', font=normal(9))
        set_cell(ws, sr, B+6, 'આચાર્યની સહી',      font=normal(9))
        return sr + 2

    current_row = 1
    i = 0
    while i < len(students):
        left_s  = students[i]
        right_s = students[i+1] if i+1 < len(students) else None
        write_student_block(ws_m, left_s,  2,  current_row, subjects)
        if right_s:
            write_student_block(ws_m, right_s, 12, current_row, subjects)
        block_height = 10 + 1 + len(subjects) + 4
        current_row += block_height + 2
        i += 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
